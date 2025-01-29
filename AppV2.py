import streamlit as st
import pandas as pd
import numpy as np
import numpy_financial as npf
import plotly.graph_objects as go
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Page Title
st.title("NBA Team Underwriting Dashboard")

# Team Selection
team = st.selectbox("Select Team", options=["Select Team", "Memphis Grizzlies"], index=0)

# Generate quarters for dropdown (still used for entry/exit periods)
def generate_quarters(start_year, end_year):
    quarters = []
    for year in range(start_year, end_year + 1):
        for quarter in range(1, 5):
            quarters.append(f"{quarter}Q{year % 100:02d}")
    return quarters

quarter_options = generate_quarters(2025, 2040)

# Team-specific inputs and calculations
if team == "Memphis Grizzlies":
    st.sidebar.header("Inputs")
    starting_revenue = 220
    starting_debt = 300
    ending_debt = 250
    starting_enterprise_value = 2112

    ownership_stake = st.sidebar.number_input("Desired Ownership Stake (%)", min_value=1.0, value=5.0, step=0.5)
    starting_enterprise_value = st.sidebar.number_input(
        "Starting Enterprise Value ($M)", min_value=0.0, value=float(starting_enterprise_value), step=10.0
    )
    starting_debt = st.sidebar.number_input(
        "Starting Debt ($M)", min_value=0.0, value=float(starting_debt), step=5.0, max_value=475.0
    )
    ending_debt = st.sidebar.number_input(
        "Ending Debt ($M)", min_value=0.0, value=float(ending_debt), step=5.0, max_value=475.0
    )

    debt_paid = starting_debt - ending_debt
    starting_equity = starting_enterprise_value - starting_debt  # Updated entry equity calculation

    starting_revenue = st.sidebar.number_input(
        "Starting Revenue ($M)", min_value=0.0, value=float(starting_revenue), step=10.0
    )

    st.sidebar.write(f"Starting Equity: ${(starting_equity):.0f}M")
    entry_quarter = st.sidebar.selectbox("Entry Quarter", options=quarter_options, index=quarter_options.index("2Q25"))
    exit_quarter = st.sidebar.selectbox("Exit Quarter", options=quarter_options, index=quarter_options.index("2Q32"))
    desired_revenue_growth = st.sidebar.number_input("Desired Revenue Growth (%)", min_value=0.0, value=10.0, step=0.1)

    # Convert Quarter to Year
    def quarter_to_year(quarter):
        quarter_num = int(quarter[0])
        year = int("20" + quarter[2:])
        return year + (quarter_num - 1) / 4

    entry_year = int(quarter_to_year(entry_quarter))
    exit_year = int(quarter_to_year(exit_quarter))
    holding_period_years = exit_year - entry_year

    # Initial TEV/Revenue multiple
    entry_tev_revenue = starting_enterprise_value / starting_revenue

    # Initialize new_tev_revenue in session state if it doesn't exist
    if "new_tev_revenue" not in st.session_state:
        st.session_state.new_tev_revenue = entry_tev_revenue  # Default to entry TEV/Revenue multiple

    # TEV/Revenue Multiple Reset Options
    def reset_to_average():
        st.session_state.new_tev_revenue = 11.9

    def reset_to_entry():
        st.session_state.new_tev_revenue = entry_tev_revenue

    def reset_to_comps():
        st.session_state.new_tev_revenue = 10.4

    # Buttons to set specific multiples
    col_set_buttons = st.columns([1, 1, 1], gap="small")
    with col_set_buttons[0]:
        if st.button("Entry Multiple"):
            reset_to_entry()
    with col_set_buttons[1]:
        if st.button("League Avg Multiple"):
            reset_to_average()
    with col_set_buttons[2]:
        if st.button("Closest Comps Multiple"):
            reset_to_comps()

    # Slider for adjusting the exit multiple
    st.header("EV/Revenue Multiple Scale")
    st.session_state.new_tev_revenue = st.slider(
        "Adjust Exit EV/Revenue Multiple",
        min_value=5.0,
        max_value=20.0,
        value=st.session_state.new_tev_revenue,  # Use the session state value
        step=0.1,
    )

    # Graph 1: TEV/Revenue Comparison
    fig_tev_revenue = go.Figure()
    fig_tev_revenue.add_trace(go.Bar(
        x=["Entry", "NBA Average", "Comps", "Exit"],
        y=[
            entry_tev_revenue,
            11.9,  # League Avg Multiple
            10.4,  # Closest Comps Multiple
            st.session_state.new_tev_revenue,  # Exit Multiple from session state
        ],
        marker_color=["#5D76A9", "gray", "darkgrey", "#5D76A9"],
        text=[
            f"{entry_tev_revenue:.1f}",
            "11.9",
            "10.4",
            f"{st.session_state.new_tev_revenue:.1f}",
        ],
        textposition="outside"
    ))
    fig_tev_revenue.update_layout(
        title="EV/Revenue Comparison",
        yaxis_title="EV/Revenue Multiple",
        template="plotly_white",
        barmode="group",
        height=500,
        width=600
    )

    # Graph 2: Revenue Growth Comparison
    comparables = {"Hornets": 14, "Hawks": 12, "Pelicans": 9}
    fig_growth = go.Figure()
    fig_growth.add_trace(go.Bar(
        x=list(comparables.keys()) + ["Imputed"],
        y=list(comparables.values()) + [desired_revenue_growth],
        marker_color=["#1D1160", "#C8102E", "#0C2340", "#5D76A9"],
        text=[f"{val:.1f}" for val in list(comparables.values()) + [desired_revenue_growth]],
        textposition="outside"
    ))
    fig_growth.update_layout(
        title="Desired Revenue Growth Rate vs Comparables",
        yaxis_title="Revenue Growth Rate (%)",
        template="plotly_white",
        barmode="group",
        height=500,
        width=600
    )

    # Display Graphs
    col1, col2 = st.columns(2)
    with col1:
        st.plotly_chart(fig_tev_revenue, use_container_width=True)
    with col2:
        st.plotly_chart(fig_growth, use_container_width=True)



    # Revenue Projections Based on Desired Revenue Growth
    annual_revenue_growth = desired_revenue_growth / 100
    projected_revenue = [
        starting_revenue * ((1 + annual_revenue_growth) ** year)
        for year in range(int(holding_period_years) + 1)
    ]

    # Corrected Entry & Exit Cash Flow Calculations
    entry_cash_flow = ownership_stake / 100 * starting_equity  # Ownership % of entry equity
    exit_equity = (projected_revenue[-1] * st.session_state.new_tev_revenue) - ending_debt  # Exit equity = TEV - Ending Debt
    exit_cash_flow = ownership_stake / 100 * exit_equity  # Ownership % of exit equity

    cash_flows = [-entry_cash_flow] + [0] * (int(holding_period_years) - 1) + [exit_cash_flow]

    def calculate_irr(cash_flows):
        return npf.irr(cash_flows) * 100

    irr = calculate_irr(cash_flows)
    moic = exit_cash_flow / abs(entry_cash_flow)

    # Projections Table
    years = list(range(entry_year, exit_year + 1))
    projections_table = pd.DataFrame({
        "": ["Revenue", "Debt Level", "Cash Flow"],
        **{
            str(year): [
                projected_revenue[i],  # Revenue
                starting_debt - (debt_paid * i / (len(years) - 1)),  # Debt Paydown
                cash_flows[i]
            ]
            for i, year in enumerate(years)
        }
    })


    # Styled HTML Table
    def generate_styled_table_horizontal(df):
        table_html = '<table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;">'
        table_html += '<thead><tr style="background-color: #0056b3; color: white; font-weight: bold; text-align: center;">'
        table_html += ''.join(f'<th style="padding: 10px; border: 1px solid #ddd;">{col}</th>' for col in df.columns)
        table_html += '</tr></thead><tbody>'

        for _, row in df.iterrows():
            table_html += f'<tr style="background-color: white; text-align: center;">'
            for i, cell in enumerate(row):
                # Check if the cell is numeric
                if isinstance(cell, (float, int)):
                    # Format negative numbers with parentheses
                    formatted_cell = f"({abs(cell):,.1f})" if cell < 0 else f"{cell:,.1f}"
                else:
                    # Leave non-numeric cells as-is
                    formatted_cell = cell
                # Bold the first column (row labels)
                table_html += f'<td style="padding: 10px; border: 1px solid #ddd; {"font-weight: bold;" if i == 0 else ""}">{formatted_cell}</td>'
            table_html += '</tr>'

        table_html += '</tbody></table>'
        return table_html


    st.subheader("Projected Financials (Annual, in $M)")
    st.markdown(generate_styled_table_horizontal(projections_table), unsafe_allow_html=True)

    st.subheader("Investment Summary (in $M)")
        
    investment_summary = pd.DataFrame({
        "Metric": [
            "Ownership Stake",
            "IRR (%)",
            "MOIC",
            "Entry Equity",
            "Exit Equity",
            "Debt Paid Off",
            "Entry Multiple",
            "Exit Multiple",
            "Revenue Growth Rate (%)",
            "Holding Period"
        ],
        "Value": [
            f"{ownership_stake:.1f}%",
            f"{irr:.1f}%",
            f"{moic:.1f}x",
            f"${entry_cash_flow:,.0f}",
            f"${exit_cash_flow:,.0f}",
            f"${debt_paid:,.0f}",
            f"{entry_tev_revenue:.1f}x",
            f"{st.session_state.new_tev_revenue:.1f}x",  # Updated
            f"{desired_revenue_growth:.1f}%",
            f"{holding_period_years:.0f}yrs"
        ]
    })

    
    def generate_summary_table_html(df):
        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<thead><tr style="background-color: #0056b3; color: white; text-align: left;">'
        html += ''.join(f'<th style="padding: 8px; border: 1px solid #ddd;">{col}</th>' for col in df.columns)
        html += '</tr></thead><tbody>'
        for _, row in df.iterrows():
            html += '<tr>'
            for col_idx, cell in enumerate(row):
                if col_idx == 1:  # Right-align the second column (Value column)
                    html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{cell}</td>'
                else:  # Left-align the first column (Metric column)
                    html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: left;">{cell}</td>'
            html += '</tr>'
        html += '</tbody></table>'
        return html


    # Display the table
    st.markdown(generate_summary_table_html(investment_summary), unsafe_allow_html=True)

    # Function to style Excel headers
    def style_headers(ws, start_row, start_col, end_col, underline=False, bold=True, empty_col=None):
        header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=bold, underline='single' if underline else None)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Handle the optional empty column
        if empty_col is not None:
            empty_cell = ws.cell(row=start_row, column=empty_col)
            empty_cell.fill = header_fill
            empty_cell.font = Font(color="FFFFFF")
            empty_cell.value = None

    def export_to_excel(projections_df, summary_df):
        # Create a workbook
        wb = Workbook()

        # Active sheet: Investment Summary
        ws_summary = wb.active
        ws_summary.title = "Investment Summary"

        # Add Projections Table
        for row_idx, row in enumerate(dataframe_to_rows(projections_df, index=False, header=True), start=2):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        # Style Headers
        style_headers(ws_summary, start_row=2, start_col=2, end_col=len(projections_df.columns) + 1, underline=True, bold=True)

        ## ---- DYNAMIC REVENUE ROW ---- ##
        growth_rate_cell = "$C$18"  # TargetCo revenue growth rate
        revenue_row = 3  # Revenue is in row 3

        for col_idx in range(3, len(projections_df.columns) + 2):  # Start at C3
            if col_idx == 3:
                ws_summary.cell(row=revenue_row, column=col_idx, value=projections_df.iloc[0, 1])
            else:
                prev_col = ws_summary.cell(row=revenue_row, column=col_idx - 1).coordinate
                current_cell = ws_summary.cell(row=revenue_row, column=col_idx)
                current_cell.value = f"={prev_col} * (1 + {growth_rate_cell})"
                current_cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        ## ---- DYNAMIC DEBT LEVEL ROW ---- ##
        debt_row = 4  # Debt Level is in row 4
        debt_paid_off_cell = f"$C$15/LEFT($C$19, LEN($C$19)-3)"

        for col_idx in range(3, len(projections_df.columns) + 2):  # Start at C4
            if col_idx == 3:
                ws_summary.cell(row=debt_row, column=col_idx, value=projections_df.iloc[1, 1])
            else:
                prev_col = ws_summary.cell(row=debt_row, column=col_idx - 1).coordinate
                current_cell = ws_summary.cell(row=debt_row, column=col_idx)
                current_cell.value = f"={prev_col} - {debt_paid_off_cell}"
                current_cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        ## ---- ADD UNDERLINE BELOW CASH FLOW ---- ##
        underline_row = 5  # Row below Cash Flow
        for col_idx in range(2, len(projections_df.columns) + 2):  
            cell = ws_summary.cell(row=underline_row, column=col_idx)
            cell.border = Border(bottom=Side(style="thin"))  # Adds underline

        ## ---- ENTERPRISE VALUE ROW ---- ##
        ev_row = 6  # Enterprise Value is in row 6
        ws_summary.cell(row=ev_row, column=2, value="Enterprise Value")  # B6 Label

        # Correctly identify the first revenue column coordinate (C3)
        first_col = ws_summary.cell(row=revenue_row, column=3).coordinate  # C3 (First Year Revenue)

        # Exit EV should be placed exactly `holding_period_years` columns to the right of first_col (C3)
        exit_col_index = 3 + holding_period_years  # This ensures the exit EV formula appears in the correct last year column
        last_col = ws_summary.cell(row=revenue_row, column=exit_col_index).coordinate  # Last Year Revenue cell

        # Place EV formulas in the correct locations
        entry_ev_cell = ws_summary.cell(row=ev_row, column=3)
        entry_ev_cell.value = f"={first_col} * LEFT($C$16,LEN($C$16)-1)"  # Entry EV Formula in First Year (C6)
        entry_ev_cell.number_format = "$#,##0.0"  # Format as currency with 1 decimal

        exit_ev_cell = ws_summary.cell(row=ev_row, column=exit_col_index)
        exit_ev_cell.value = f"={last_col} * LEFT($C$17,LEN($C$17)-1)"  # Exit EV Formula in correct final column
        exit_ev_cell.number_format = "$#,##0.0"  # Format as currency with 1 decimal

        # Leave Intermediate Columns Blank
        for col_idx in range(4, exit_col_index):  # Blank out D6 to second-to-last column before exit EV
            ws_summary.cell(row=ev_row, column=col_idx, value="")

        # Ensure gridlines are hidden
        ws_summary.sheet_view.showGridLines = False

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 1
        ws_summary.column_dimensions["B"].width = 20
        for col_idx in range(2, len(projections_df.columns) + 3):
            ws_summary.column_dimensions[chr(64 + col_idx)].width = 15

        # **Add Summary Table Below the Projections Table**
        summary_start_row = len(projections_df) + 6  # Dynamic placement

        for row_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=summary_start_row):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        # Style Summary Table headers
        style_headers(ws_summary, start_row=summary_start_row, start_col=2, end_col=3)

                # Overwrite C11: IRR of Cash Flow for Holding Period
        cash_flow_range = f"C{revenue_row + 2}:J{revenue_row + 2}"  # Adjusted for the row where cash flows are stored
        ws_summary.cell(row=11, column=3).value = f"=IRR({cash_flow_range})"
        ws_summary.cell(row=11, column=3).number_format = "0.0%"

        # Overwrite C12: MOIC Calculation
        ws_summary.cell(row=12, column=3).value = "=C14/C13"
        ws_summary.cell(row=12, column=3).number_format = "0.0x"

        # Overwrite C13: Equity Value at Entry
        ws_summary.cell(row=13, column=3).value = "=(C6-C4)*C10"
        ws_summary.cell(row=13, column=3).number_format = "$#,##0.0"

        # Overwrite C14: Equity Value at Exit
        exit_ev_col = 3 + holding_period_years  # Dynamically calculate the column index for exit EV
        exit_ev_cell = ws_summary.cell(row=ev_row, column=exit_ev_col).coordinate  # J6 or last column of EV
        exit_debt_col = ws_summary.cell(row=debt_row, column=exit_ev_col).coordinate  # J4 or last column of debt
        ws_summary.cell(row=14, column=3).value = f"=({exit_ev_cell}-{exit_debt_col})*C10"
        ws_summary.cell(row=14, column=3).number_format = "$#,##0.0"

        for row_idx in range(10, 20):  # Rows 10 through 19
            cell = ws_summary.cell(row=row_idx, column=3)  # Column C
            cell.alignment = Alignment(horizontal="right", vertical="center")

        return wb


    # Function to add Comparable Transactions sheet
    def add_comparable_transactions_sheet(wb, starting_enterprise_value, entry_tev_revenue, desired_revenue_growth):

        # Add Comparable Transactions Sheet
        ws_comparables = wb.create_sheet(title="Comparable Transactions")

        # Hardcoded data
        comparables_data = [
            ["Date", "Team", "Transaction Value", "TEV/Revenue", "5-Year Revenue Growth"],
            ["7/3/2024", "Charlotte Hornets", 1850, 9.1, .14],
            ["2/24/2024", "Atlanta Hawks", 2210, 11.3, .12],
            ["12/2/2023", "New Orleans Pelicans", 1980, 10.8, .09],
            ["", "TargetCo", starting_enterprise_value, entry_tev_revenue, (desired_revenue_growth/100)],
        ]

        # Add data to Comparable Transactions starting at B2
        for row_idx, row in enumerate(comparables_data, start=2):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_comparables.cell(row=row_idx, column=col_idx, value=value)
                # Format numbers based on column type
                if isinstance(value, (int, float)):
                    if col_idx == 4:  # Format Transaction Value as currency with 1 decimal
                        cell.number_format = "$#,##0.0"
                    elif col_idx == 5:  # Format EV/Sales as ##.#x
                        cell.number_format = "0.0x"
                    elif col_idx == 6:  # Format Revenue Growth as percentage
                        cell.number_format = "0.0%"
                # Center-align all information in cells
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Apply light green color to TargetCo row
                if row_idx == 6:  # TargetCo row is at index 6 (row after Lakers)
                    cell.fill = PatternFill(start_color="5D76A9", end_color="5D76A9", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
              
        # Style headers (B2:F2)
        header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, underline="single")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(2, 7):
            cell = ws_comparables.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add line under the header row (B5:F5)
        for col in range(2, 7):
            cell = ws_comparables.cell(row=5, column=col)
            cell.border = Border(bottom=Side(style="thin"))

        # Add line under the Mean row (B8:F8)
        for col in range(2, 7):
            cell = ws_comparables.cell(row=6, column=col)
            cell.border = Border(bottom=Side(style="thin"))
  

        # Add Median row dynamically
        median_row = ["Median", ""]
        median_row_formulas = [
            f"=MEDIAN(D3:D5)",  # Transaction Value
            f"=MEDIAN(E3:E5)",  # EV/Sales
            f"=MEDIAN(F3:F5)"   # Revenue Growth
        ]

        for col_idx, value in enumerate(median_row + median_row_formulas, start=2):
            cell = ws_comparables.cell(row=7, column=col_idx, value=value if col_idx <= 3 else None)
            if col_idx > 3:
                cell.value = median_row_formulas[col_idx - 4]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            if col_idx == 4:
                cell.number_format = "$#,##0.0"
            elif col_idx == 5:
                cell.number_format = "0.0x"
            elif col_idx == 6:
                cell.number_format = "0.0%"

        # Add Mean row dynamically
        mean_row = ["Mean", ""]
        mean_row_formulas = [
            f"=AVERAGE(D3:D5)",  # Transaction Value
            f"=AVERAGE(E3:E5)",  # EV/Sales
            f"=AVERAGE(F3:F5)"   # Revenue Growth
        ]

        for col_idx, value in enumerate(mean_row + mean_row_formulas, start=2):
            cell = ws_comparables.cell(row=8, column=col_idx, value=value if col_idx <= 3 else None)
            if col_idx > 3:
                cell.value = mean_row_formulas[col_idx - 4]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            if col_idx == 4:
                cell.number_format = "$#,##0.0"
            elif col_idx == 5:
                cell.number_format = "0.0x"
            elif col_idx == 6:
                cell.number_format = "0.0%"

        # Ensure gridlines are not shown
        ws_comparables.sheet_view.showGridLines = False

        ws_comparables.column_dimensions["A"].width = 1
        ws_comparables.column_dimensions["B"].width = 20
        ws_comparables.column_dimensions["C"].width = 20
        ws_comparables.column_dimensions["D"].width = 20
        ws_comparables.column_dimensions["E"].width = 20
        ws_comparables.column_dimensions["F"].width = 20

        return wb

    # Export Button in Streamlit
    def export_excel_button(projections_table, investment_summary, starting_enterprise_value, entry_tev_revenue, desired_revenue_growth):
        wb = export_to_excel(projections_table, investment_summary)
        wb = add_comparable_transactions_sheet(wb, starting_enterprise_value, entry_tev_revenue, desired_revenue_growth)

        # Save Workbook to BytesIO for Download
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Export Button in Streamlit
        st.download_button(
            label="Download Excel File",
            data=output,
            file_name="NBA_Team_Underwriting.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Call the function
    export_excel_button(projections_table, investment_summary, starting_enterprise_value, entry_tev_revenue, desired_revenue_growth)


