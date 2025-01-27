import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import numpy_financial as npf
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Alignment
from statistics import median


# Page Title
st.title("NBA Team Underwriting Dashboard")

def generate_quarters(start_year, end_year):
    quarters = []
    for year in range(start_year, end_year + 1):
        for quarter in range(1, 5):
            quarters.append(f"{quarter}Q{year % 100:02d}")
    return quarters

quarter_options = generate_quarters(2025, 2040)

# Team Selection Dropdown
team = st.selectbox("Select Team", options=["Select Team", "Boston Celtics"], index=0)

# Set financial inputs based on the selected team
if team == "Boston Celtics":
    starting_revenue = 390
    starting_debt = 325
    starting_enterprise_value = 5660

    st.sidebar.header("Inputs")

    ownership_stake = st.sidebar.number_input("Desired Ownership Stake (%)", min_value=1.0, value=5.0, step=0.5)
    starting_enterprise_value = st.sidebar.number_input(
        "Starting Enterprise Value ($M)", min_value=0.0, value=float(starting_enterprise_value), step=10.0
    )
    starting_debt = st.sidebar.number_input(
        "Starting Debt ($M)", min_value=0.0, value=float(starting_debt), step=5.0
    )
    starting_equity = starting_enterprise_value - starting_debt
    starting_revenue = st.sidebar.number_input(
        "Starting Revenue ($M)", min_value=0.0, value=float(starting_revenue), step=10.0
    )

    st.sidebar.text(f"Starting TEV: ${starting_enterprise_value:.0f}M")

    entry_quarter = st.sidebar.selectbox("Entry Quarter", options=quarter_options, index=quarter_options.index("2Q25"))
    exit_quarter = st.sidebar.selectbox("Exit Quarter", options=quarter_options, index=quarter_options.index("2Q32"))
    desired_moic = st.sidebar.number_input("Desired MOIC (x)", min_value=1.0, value=2.5, step=0.1)

    def quarter_to_year(quarter):
        try:
            quarter_num = int(quarter[0])
            if quarter_num not in [1, 2, 3, 4]:
                raise ValueError("Quarter must be between 1 and 4.")
            year = int("20" + quarter[2:])
            return year + (quarter_num - 1) / 4
        except (ValueError, IndexError):
            st.error("Invalid quarter format. Use the format '1Q25'.")
        return None

    entry_year = quarter_to_year(entry_quarter)
    exit_year = quarter_to_year(exit_quarter)
    holding_period_years = exit_year - entry_year

    # Initial TEV/Revenue multiple
    entry_tev_revenue = starting_enterprise_value / starting_revenue

    def reset_to_average():
        return 11.9

    def reset_to_entry():
        return entry_tev_revenue

    def reset_to_comps():
        return 13.1

    new_tev_revenue = entry_tev_revenue
    
    # Adjust layout with minimal gap
    col_set_buttons = st.columns([1, 1, 1], gap="small")

    with col_set_buttons[0]:
        if st.button("Entry Multiple"):
            new_tev_revenue = reset_to_entry()

    with col_set_buttons[1]:
        if st.button("League Avg Multiple"):
            new_tev_revenue = reset_to_average()

    with col_set_buttons[2]:
        if st.button("Closest Comps Multiple"):
            new_tev_revenue = reset_to_comps()

        
    st.header("TEV/Revenue Multiple Scale")
    new_tev_revenue = st.slider("Adjust Exit TEV/Revenue Multiple", min_value=5.0, max_value=20.0, value=new_tev_revenue, step=0.1)

    # Graph 1
    fig_tev_revenue = go.Figure()
    fig_tev_revenue.add_trace(go.Bar(
        x=["Entry", "NBA Average", "Comps", "Exit"],
        y=[entry_tev_revenue, 11.9, 13.1, new_tev_revenue],
        marker_color=["#007A33", "gray", "navy", "#007A33"],
        text=[f"{entry_tev_revenue:.1f}", "11.9", "13.1", f"{new_tev_revenue:.1f}"],
        textposition="outside"
    ))
    fig_tev_revenue.update_layout(
        title="TEV/Revenue Comparison",
        yaxis_title="TEV/Revenue Multiple",
        template="plotly_white",
        barmode="group",
        height=500,
        width=600
    )

    def solve_for_revenue_growth():
        def objective(revenue_growth):
            exit_revenue = starting_revenue * ((1 + revenue_growth / 100) ** holding_period_years)
            exit_tev = exit_revenue * new_tev_revenue
            moic = exit_tev / starting_enterprise_value
            return abs(moic - desired_moic)

        from scipy.optimize import minimize
        result = minimize(objective, x0=[10], bounds=[(0, 30)])
        return result.x[0]

    required_revenue_growth = solve_for_revenue_growth()

    # Graph 2
    comparables = {"Warriors": 14, "Knicks": 12, "Lakers": 9}
    fig_growth = go.Figure()
    fig_growth.add_trace(go.Bar(
        x=list(comparables.keys()) + ["Required"],
        y=list(comparables.values()) + [required_revenue_growth],
        marker_color=["#FFC72C", "#006BB6", "#552583", "#007A33"],
        text=[f"{val:.1f}" for val in list(comparables.values()) + [required_revenue_growth]],
        textposition="outside"
    ))
    fig_growth.update_layout(
        title="Implied Revenue Growth Rate vs Comparables",
        yaxis_title="Revenue Growth Rate (%)",
        template="plotly_white",
        barmode="group",
        height=500,
        width=600
    )

    col1, col2 = st.columns(2)
    with col1:
        st.plotly_chart(fig_tev_revenue, use_container_width=True)
    with col2:
        st.plotly_chart(fig_growth, use_container_width=True)

    projected_revenue = [starting_revenue * ((1 + required_revenue_growth / 100) ** year) for year in range(int(holding_period_years) + 1)]
    entry_cash_flow = ownership_stake / 100 * starting_enterprise_value
    exit_cash_flow = ownership_stake / 100 * (projected_revenue[-1] * new_tev_revenue)

    cash_flows = [
        -entry_cash_flow
    ] + [0] * (int(holding_period_years) - 1) + [exit_cash_flow]

    def calculate_irr(cash_flows):
        return npf.irr(cash_flows) * 100

    irr = calculate_irr(cash_flows)
    moic = exit_cash_flow / abs(entry_cash_flow)

    def generate_styled_table(df):
        table_html = '<table style="border-collapse: collapse; width: 100%; font-family: Arial, sans-serif;">'
        table_html += '<tr style="background-color: #0056b3; color: white; font-weight: bold; text-align: center;">'
        for col in df.columns:
            table_html += f'<th style="padding: 10px; border: 1px solid #ddd;">{col}</th>'
        table_html += '</tr>'
        for _, row in df.iterrows():
            table_html += '<tr style="background-color: white; text-align: center;">'
            for cell in row:
                # Format numbers to 1 decimal place and handle negatives with parentheses
                if isinstance(cell, (float, int)):
                    formatted_cell = f"({abs(cell):,.1f})" if cell < 0 else f"{cell:,.1f}"
                else:
                    formatted_cell = cell  # Keep non-numeric cells as is
                table_html += f'<td style="padding: 10px; border: 1px solid #ddd;">{formatted_cell}</td>'
            table_html += '</tr>'
        table_html += '</table>'
        return table_html

    projections_styled = pd.DataFrame({
        " ": ["Revenue", "Cash Flow"],
        **{f"{int(entry_year) + i}": [projected_revenue[i], cash_flows[i]] for i in range(len(projected_revenue))}
    })


    # Generate the styled HTML table
    html_table = generate_styled_table(projections_styled)

    def generate_quarters_table_html_horizontal(df):
        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<thead><tr style="background-color: #0056b3; color: white; text-align: center;">'
        html += ''.join(f'<th style="padding: 8px; border: 1px solid #ddd;">{col}</th>' for col in df.columns)
        html += '</tr></thead><tbody>'
        for _, row in df.iterrows():
            html += '<tr>'
            for cell in row:
                # Format numbers to 1 decimal place and handle negatives with parentheses
                if isinstance(cell, (float, int)):
                    formatted_cell = f"({abs(cell):,.1f})" if cell < 0 else f"{cell:,.1f}"
                else:
                    formatted_cell = cell  # Keep non-numeric cells as is
                html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{formatted_cell}</td>'
            html += '</tr>'
        html += '</tbody></table>'
        return html

    # Toggle for displaying quarterly data
    st.subheader("Projected Financials (in $MM)")
    # Add a dropdown to toggle between "Years" and "Quarters"
    view_option = st.selectbox("View By", options=["Years", "Quarters"], index=0)

    if view_option == "Quarters":
        # Generate quarters data for the expanded view
        quarter_labels = []
        revenue_row = []
        cash_flow_row = []

        for year_idx, annual_revenue in enumerate(projected_revenue):
            year = int(entry_year) + year_idx
            for q in range(1, 5):
                quarter_label = f"{q}Q{str(year)[-2:]}"
                quarter_labels.append(quarter_label)

                # Divide annual revenue by 4 for each quarter (keep as numeric)
                revenue_row.append(annual_revenue / 4)

                # Cash flow logic (keep as numeric)
                if quarter_label == entry_quarter:
                    # Entry Quarter: Investment (negative cash flow)
                    cash_flow_row.append(cash_flows[0])
                elif quarter_label == exit_quarter:
                    # Exit Quarter: Return (positive cash flow)
                    cash_flow_row.append(cash_flows[-1])
                else:
                    # Intermediate Quarters: No cash flow
                    cash_flow_row.append(0.0)

        # Create a DataFrame for the quarters table
        quarters_table = pd.DataFrame(
            [revenue_row, cash_flow_row],
            columns=quarter_labels,
            index=["Revenue", "Cash Flow"]
        ).reset_index()

        quarters_table.rename(columns={"index": ""}, inplace=True)

        # Generate the styled HTML table
        def generate_quarters_table_html_horizontal(df):
            html = '<table style="width: 100%; border-collapse: collapse;">'
            html += '<thead><tr style="background-color: #0056b3; color: white; text-align: center;">'
            html += ''.join(f'<th style="padding: 8px; border: 1px solid #ddd;">{col}</th>' for col in df.columns)
            html += '</tr></thead><tbody>'
            for _, row in df.iterrows():
                html += '<tr>'
                for cell in row:
                    # Format numbers to 1 decimal place and handle negatives with parentheses
                    if isinstance(cell, (float, int)):
                        formatted_cell = f"({abs(cell):,.1f})" if cell < 0 else f"{cell:,.1f}"
                    else:
                        formatted_cell = cell  # Keep non-numeric cells as is
                    html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{formatted_cell}</td>'
                html += '</tr>'
            html += '</tbody></table>'
            return html

        # Generate and display the styled HTML table
        html_quarters_table = generate_quarters_table_html_horizontal(quarters_table)
        st.markdown(html_quarters_table, unsafe_allow_html=True)

    else:
        # Show the original yearly table
        st.markdown(html_table, unsafe_allow_html=True)


    st.subheader("Investment Summary (in $MM)")
        
    investment_summary = pd.DataFrame({
        "Metric": [
            "Entry Equity",
            "Exit Equity",
            "IRR (%)",
            "MOIC",
            "Entry Multiple",
            "Exit Multiple",
            "Implied Revenue Growth Rate (%)"
        ],
        "Value": [
            f"${entry_cash_flow:,.0f}",
            f"${exit_cash_flow:,.0f}",
            f"{irr:.1f}%",
            f"{moic:.1f}x",
            f"{entry_tev_revenue:.1f}x",
            f"{new_tev_revenue:.1f}x",
            f"{required_revenue_growth:.1f}%"
        ]
    })
    
    def generate_summary_table_html(df):
        html = '<table style="width: 100%; border-collapse: collapse;">'
        html += '<thead><tr style="background-color: #0056b3; color: white; text-align: left;">'
        html += ''.join(f'<th style="padding: 8px; border: 1px solid #ddd;">{col}</th>' for col in df.columns)
        html += '</tr></thead><tbody>'
        for _, row in df.iterrows():
            html += '<tr>'
            for col_idx, cell in enumerate(row):
                if col_idx == 1:  # Right-align the second column (Value column)
                    html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{cell}</td>'
                else:  # Left-align the first column (Metric column)
                    html += f'<td style="padding: 8px; border: 1px solid #ddd; text-align: left;">{cell}</td>'
            html += '</tr>'
        html += '</tbody></table>'
        return html


    # Display the table
    st.markdown(generate_summary_table_html(investment_summary), unsafe_allow_html=True)


    # Function to style Excel headers
    def style_headers(ws, start_row, start_col, end_col, underline=False, bold=True, empty_col=None):
        header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=bold, underline='single' if underline else None)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Handle the optional empty column
        if empty_col is not None:
            empty_cell = ws.cell(row=start_row, column=empty_col)
            empty_cell.fill = header_fill
            empty_cell.font = Font(color="FFFFFF")  # Make it consistent but empty
            empty_cell.value = None

    def process_cell_as_numeric_with_format(ws, row, col):
        # Get the current cell value
        cell_value = ws.cell(row=row, column=col).value

        # Convert cell value to numeric if it contains "x"
        if isinstance(cell_value, str) and "x" in cell_value:
            numeric_value = float(cell_value.replace("x", ""))
        else:
            numeric_value = cell_value

        # Update the cell with the numeric value
        ws.cell(row=row, column=col).value = numeric_value

        # Apply the "0.0x" number format to the cell
        ws.cell(row=row, column=col).number_format = "0.0x"

        return numeric_value


    def export_to_excel_one_sheet(projections_df, summary_df):
        # Create a workbook
        wb = Workbook()

        # Active sheet: Investment Summary
        ws_summary = wb.active
        ws_summary.title = "Investment Summary"

        # Add Projections Table starting at B2
        for row_idx, row in enumerate(dataframe_to_rows(projections_df, index=False, header=True), start=2):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                # Format numbers in currency with parentheses for negatives
                if isinstance(value, (int, float)):
                    cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        # Style Projections Table headers
        style_headers(ws_summary, start_row=2, start_col=2, end_col=len(projections_df.columns) + 1, underline=True, bold=True)

        # Add Summary Table starting dynamically below Projections Table
        summary_start_row = len(projections_df) + 5
        for row_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=summary_start_row):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                # Format numbers in currency with parentheses for negatives
                if isinstance(value, (int, float)):
                    cell.number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        # Style Summary Table headers
        style_headers(ws_summary, start_row=summary_start_row, start_col=2, end_col=3)

        for row in range(summary_start_row + 1, summary_start_row + len(summary_df) + 1):
            ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")

        # Add dynamic IRR and MOIC formulas to the Summary Table
        cash_flow_range = f"'Investment Summary'!C4:J4"  # Fixed to span columns C to J in row 4
        ws_summary.cell(row=summary_start_row + len(summary_df), column=2).value = "IRR (%)"
        irr_cell = ws_summary.cell(row=summary_start_row + len(summary_df), column=3)
        irr_cell.value = f"=IRR({cash_flow_range})"
        irr_cell.number_format = "0.0%"  # Automatically formats as a percentage with 1 decimal place

        # MOIC formula
        entry_cash_flow_cell = ws_summary.cell(row=4, column=3).coordinate  # Entry cash flow in Projections
        exit_cash_flow_cell = ws_summary.cell(row=4, column=10).coordinate  # Exit cash flow in Projections
        ws_summary.cell(row=summary_start_row + len(summary_df) + 1, column=2).value = "MOIC (x)"
        moic_cell = ws_summary.cell(row=summary_start_row + len(summary_df) + 1, column=3)
        moic_cell.value = f"={exit_cash_flow_cell}/ABS({entry_cash_flow_cell})"
        moic_cell.number_format = "0.0x"

        ws_summary.delete_rows(idx=10, amount=2)

        # Process cells and retrieve numeric values
        c10_value = process_cell_as_numeric_with_format(ws_summary, 10, 3)
        c11_value = process_cell_as_numeric_with_format(ws_summary, 11, 3)

        # Add data to row 5 starting at B5
        ws_summary.cell(row=5, column=2).value = "Enterprise Value"  # B5
        ws_summary.cell(row=5, column=3).value = f"=C10*C3"          # C5
        ws_summary.cell(row=5, column=3).number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"
        for col in range(4, 10):                             # D5 to I5
            ws_summary.cell(row=5, column=col).value = ""
        ws_summary.cell(row=5, column=10).value = f"=C11*J3"         # J5
        ws_summary.cell(row=5, column=10).number_format = "_($#,##0.0_);_($(#,##0.0);_($\"-\"??_);_(@_)"

        ws_summary.sheet_view.showGridLines = False

        ws_summary.column_dimensions["A"].width = 1
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 12
        ws_summary.column_dimensions["D"].width = 12
        ws_summary.column_dimensions["E"].width = 12
        ws_summary.column_dimensions["F"].width = 12
        ws_summary.column_dimensions["G"].width = 12
        ws_summary.column_dimensions["H"].width = 12
        ws_summary.column_dimensions["I"].width = 12
        ws_summary.column_dimensions["J"].width = 12
        return wb

    def add_comparable_transactions_sheet(wb):
        # Add Comparable Transactions Sheet
        ws_comparables = wb.create_sheet(title="Comparable Transactions")

        # Hardcoded data
        comparables_data = [
            ["Date", "Team", "Transaction Value", "TEV/Revenue", "5-Year Revenue Growth"],
            ["7/3/2024", "Golden State Warriors", 6250, 14.9, .14],
            ["2/24/2024", "New York Knicks", 5340, 12.8, .12],
            ["12/2/2023", "Los Angeles Lakers", 5870, 11.6, .09],
        ]

        # Add data to Comparable Transactions starting at B2
        for row_idx, row in enumerate(comparables_data, start=2):
            for col_idx, value in enumerate(row, start=2):  # Start at column B
                cell = ws_comparables.cell(row=row_idx, column=col_idx, value=value)
                # Format numbers based on column type
                if isinstance(value, (int, float)):
                    if col_idx == 4:  # Format Transaction Value as currency with 1 decimal
                        cell.number_format = "$#,##0.0"
                    elif col_idx == 5:  # Format EV/Sales as ##.#x
                        cell.number_format = "0.0x"
                    elif col_idx == 6:  # Format Revenue Growth as percentage
                        cell.number_format = "0.0%"
                # Center-align all information in cells
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style headers (B2:F2)
        header_fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, underline="single")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(2, 7):
            cell = ws_comparables.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add line under the header row (B5:F5)
        for col in range(2, 7):
            cell = ws_comparables.cell(row=5, column=col)
            cell.border = Border(bottom=Side(style="thin"))

        # Add Median row dynamically
        median_row = ["Median", ""]
        median_row_formulas = [
            f"=MEDIAN(D3:D5)",  # Transaction Value
            f"=MEDIAN(E3:E5)",  # EV/Sales
            f"=MEDIAN(F3:F5)"   # Revenue Growth
        ]

        for col_idx, value in enumerate(median_row + median_row_formulas, start=2):
            cell = ws_comparables.cell(row=6, column=col_idx, value=value if col_idx <= 3 else None)
            if col_idx > 3:
                cell.value = median_row_formulas[col_idx - 4]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            if col_idx == 4:
                cell.number_format = "$#,##0.0"
            elif col_idx == 5:
                cell.number_format = "0.0x"
            elif col_idx == 6:
                cell.number_format = "0.0%"

        # Add Mean row dynamically
        mean_row = ["Mean", ""]
        mean_row_formulas = [
            f"=AVERAGE(D3:D5)",  # Transaction Value
            f"=AVERAGE(E3:E5)",  # EV/Sales
            f"=AVERAGE(F3:F5)"   # Revenue Growth
        ]

        for col_idx, value in enumerate(mean_row + mean_row_formulas, start=2):
            cell = ws_comparables.cell(row=7, column=col_idx, value=value if col_idx <= 3 else None)
            if col_idx > 3:
                cell.value = mean_row_formulas[col_idx - 4]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            if col_idx == 4:
                cell.number_format = "$#,##0.0"
            elif col_idx == 5:
                cell.number_format = "0.0x"
            elif col_idx == 6:
                cell.number_format = "0.0%"

        # Ensure gridlines are not shown
        ws_comparables.sheet_view.showGridLines = False

        ws_comparables.column_dimensions["A"].width = 1
        ws_comparables.column_dimensions["B"].width = 20
        ws_comparables.column_dimensions["C"].width = 20
        ws_comparables.column_dimensions["D"].width = 20
        ws_comparables.column_dimensions["E"].width = 20
        ws_comparables.column_dimensions["F"].width = 20

        return wb





    # Prepare DataFrames for Export
    if view_option == "Years":
        projections_df = projections_styled  # Use the annual projections
    else:
        projections_df = pd.DataFrame(
            [revenue_row, cash_flow_row],
            columns=quarter_labels,
            index=["Revenue ($M)", "Cash Flow ($M)"]
        ).reset_index()

    summary_df = investment_summary  # Use the summary DataFrame

    # Generate Workbook
    wb = export_to_excel_one_sheet(projections_df, summary_df)
    wb = add_comparable_transactions_sheet(wb)

    # Save Workbook to BytesIO for Download
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Export Button in Streamlit
    st.download_button(
        label="Download Excel File",
        data=output,
        file_name="CelticsModel_v1.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
